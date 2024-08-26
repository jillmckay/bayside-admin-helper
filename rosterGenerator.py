#!/usr/bin/env python3

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from collections import namedtuple
import os
import argparse
import logging
import csv
import json
import math
from datetime import datetime

logging.basicConfig(format="%(levelname)s\t|| %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)


def get_leader_map(leader_map: str) -> dict:
    leader_map_fields = ["Group", "First Name", "Last Name"]
    group_leaders = {}
    with open(leader_map) as lm:
        reader = csv.DictReader(lm)
        if not reader.fieldnames == leader_map_fields:
            logger.error(f"Leader Map `{leader_map}` formatted incorrectly.\n"
                         f"Expected fields: {leader_map_fields}.\n"
                         f"Found: {reader.fieldnames}")
        for row in reader:
            if row["Group"] not in group_leaders:
                group_leaders[row["Group"]] =  []
            group_leaders[row["Group"]].append(row)
    return group_leaders


def normalize_record(record: dict) -> dict:
    logger.debug(f"Normalizing record for {record['First Name']} {record['Last Name']}")
    logger.debug("Stitching together address")
    for key in ['Street 1', 'Street 2', 'City', 'State', 'Postal Code']:
        norm = str(record[key])
        norm = '' if norm == 'nan' else norm  # remove nan values
        record[key] = norm
    # form address from fields
    record['Address'] = (
        f"{record['Street 1']}{', ' if record['Street 2'] else ''}{record['Street 2']}, "
        f"{record['City']}, {record['State']} {record['Postal Code'].split('-')[0]}"
    )
    logger.debug("Removing year from Birthdate")
    if isinstance(record['Birthdate'], str):
        record['Birthdate'] = str(record['Birthdate']).rsplit('/', 1)[0]
    else:
        record['Birthdate'] = ""
    keys_to_keep = {'First Name', 'Last Name', 'Address',
                    'Mobile', 'Email', 'Birthdate', 'Placements'}
    return {key: record[key] for key in keys_to_keep}

def load_registrants(registrants: str) -> dict:
    logger.debug(f"Loading registrants excel from `{registrants}`")
    records = pd.read_excel(registrants, sheet_name=0, header=2).to_dict(orient='records')
    logger.debug(f"{len(records)} registrants found")
    roster = {}
    for record in records:
        placement = record['Placements'] if isinstance(record['Placements'], str) else 'UNASSIGNED'
        record['Placements'] = placement
        record = normalize_record(record)
        # add record to roster based on placement
        if placement not in roster:
            roster[placement] = []
        if record not in roster[placement]:
            roster[placement].append(record)
    return roster


def title_cell(worksheet, col_span, title):
    worksheet.merge_cells(f"A1:{get_column_letter(col_span)}1")
    tc = worksheet["A1"]
    tc.value = title
    tc.font = Font(bold=True)
    tc.alignment = Alignment(horizontal='center', vertical='center')    


def at_a_glance(roster, leader_map, writer, num_chunks=4):
    cell = namedtuple("cell", "col row", defaults="")
    all_groups = []
    pivot = []
    for gid, groster in sorted(roster.items()):
        all_groups.extend(groster)
        group_name = gid.partition('-')[0].strip()
        pivot.append({
            "Leader Name": f"{leader_map[gid][0]['First Name']} {leader_map[gid][0]['Last Name']}" if gid != "UNASSIGNED" else "",
            "Group": group_name,
            "Members": len(groster)
        })
    pivot.append({"Group": "Total", "Members": len(all_groups)})
    all_groups = sorted(all_groups, key=lambda x: (x["Last Name"], x["First Name"]))
    at_a_glance_cols = ("Last Name", "First Name", "Placements")
    last, first, grp = at_a_glance_cols

    all_groups = [{last: p[last], first: p[first], grp: p[grp].partition('-')[0].strip()} for p in all_groups]
    aag_row_count = len(all_groups) + len(pivot) + 2  # add 2 blank rows before pivot table
    chunk_size = math.ceil(aag_row_count / num_chunks)
    first_row = 2
    for i in range(num_chunks):
        start_idx = i * chunk_size
        end_idx = (i + 1) * chunk_size if i < num_chunks else None
        chunk = all_groups[start_idx:end_idx]
        chunk_start = cell(col=1+(i*(len(at_a_glance_cols)+1)), row=first_row)
        chunk_end = cell(col=len(at_a_glance_cols)+(i*(len(at_a_glance_cols)+1)), row=first_row+len(chunk))
        df = pd.DataFrame(chunk)
        df.to_excel(writer, sheet_name="At a Glance", index=False,
                    startrow=chunk_start.row-1, startcol=chunk_start.col-1)
        worksheet = writer.sheets["At a Glance"]
        table_ref = f"{get_column_letter(chunk_start.col)}{chunk_start.row}:{get_column_letter(chunk_end.col)}{chunk_end.row}"
        table = Table(displayName=f"At_a_Glance_{i+1}", ref=table_ref,
                      tableStyleInfo=TableStyleInfo(name='TableStyleLight11',
                                                    showFirstColumn=False, showLastColumn=False,
                                                    showRowStripes=True, showColumnStripes=True))
        worksheet.add_table(table)
    # add pivot table relative to last at a glance table
    pivot_start = cell(col=chunk_start.col, row=chunk_end.row+2)
    pivot_end = cell(col=chunk_start.col+len(pivot[0].keys())-1, row=pivot_start.row+len(pivot))
    df = pd.DataFrame(pivot)
    df.to_excel(writer, sheet_name="At a Glance", index=False,
                startrow=pivot_start.row-1, startcol=pivot_start.col-1)
    worksheet = writer.sheets["At a Glance"]
    table_ref = f"{get_column_letter(pivot_start.col)}{pivot_start.row}:{get_column_letter(pivot_end.col)}{pivot_end.row}"
    table = Table(displayName=f"At_a_Glance_pivot", ref=table_ref, 
                  tableStyleInfo=TableStyleInfo(name='TableStyleLight11',
                                                showFirstColumn=False, showLastColumn=False,
                                                showRowStripes=True, showColumnStripes=True))
    worksheet.add_table(table)
    auto_fit_column_widths(worksheet)
    # skipped columns DHL
    for skipped in range(1, num_chunks):
        column_letter = get_column_letter(skipped * (len(at_a_glance_cols) + 1))
        worksheet.column_dimensions[column_letter].width = 2.0
    title_cell(worksheet, pivot_end.col, f"At a Glance -  Generated {datetime.now().strftime('%m%d%y')}")


def leader_roster(leader_map, roster, writer):
    first, last = "First Name", "Last Name"
    leaders = []
    for group in leader_map:
        groster = roster[group]
        for leader in leader_map[group]:
            # find leader in groster
            for member in groster:
                if member[first] == leader[first] and member[last] == leader[last]:
                    leaders.append(member)
                    logger.debug(f"Adding {member[first]} {member[last]} to Leader Roster")
                    break
            else:
                logger.warn(f"Did not find {leader[first]} {leader[last]} in group roster for {group}")
    group_roster("Leaders", leaders, writer)


def group_roster(group, roster, writer):
    column_order = ["First Name", "Last Name", "Address", "Mobile", "Email", "Birthdate", "Placements"]
    group_name = group.partition('-')[0].strip()
    df = pd.DataFrame(roster)
    # Reorder columns based on specified order
    df = df[column_order]
    # Write the DataFrame to Excel starting at row 2 so row 1 can be merged cells
    df.to_excel(writer, sheet_name=group_name, index=False, startrow=1)
    # get the worksheet to format the table
    worksheet = writer.sheets[group_name]
    # create table
    table = Table(displayName=f"{group_name.replace(' ','')}Roster", ref=worksheet.dimensions,
                  tableStyleInfo=TableStyleInfo(name='TableStyleLight12',
                                                showFirstColumn=False, showLastColumn=False,
                                                showRowStripes=True, showColumnStripes=True))
    # add table to worksheet
    worksheet.add_table(table)
    # add title row spanning whole table
    auto_fit_column_widths(worksheet)
    title_cell(worksheet, len(column_order), f"{group} - Generated {datetime.now().strftime('%m%d%y')}")


def auto_fit_column_widths_alt(worksheet):
    for column in worksheet.columns:
        column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C', ...)
        worksheet.column_dimensions[column_letter].bestFit = True
        

def auto_fit_column_widths(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C', ...)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
    
        adjusted_width = max_length + 2.0  # Add some extra padding
        worksheet.column_dimensions[column_letter].width = adjusted_width
        worksheet.column_dimensions[column_letter].bestFit = True


def valid_file(argument):
    # locatable file
    if not os.access(argument, os.R_OK):
        raise argparse.ArgumentTypeError(message="Not a readable file")
    return argument


def main():
    parser = argparse.ArgumentParser(
        prog="RosterGenerator",
        description="Process event registrant lists into group rosters",
        epilog="For additional help, contact jill.mckay@baysideonline.com"
    )
    parser.add_argument("--leader-map", "-l", action="store", required=True,
                        type=valid_file,
                        help=("CSV file with the following fields:\n"
                              "Group,First Name, Last Name"))
    parser.add_argument("--registrants", "-r", action="append", required=True,
                        type=valid_file,
                        help=("Exported Excel spread sheet from the "
                              "event registration Registrants pane.\n"
                              "If more than one registrants list is specifed, "
                              "they will be treated as an extension of the "
                              "registrants for the study"))
    parser.add_argument("--study-name", "-s", action="store", required=True,
                        help=("Bible Study Name"))
    parser.add_argument("--output-file", "-o", action="store", required=False,
                        help=("Name of file to output roster to. "
                              "Will be overwritten if already exists"))
    parser.add_argument("--at-a-glance-size", "-a", action="store", required=False,
                        type=int, default=4,
                        help="Define how many groupings the At A Glance Page has")
    parser.add_argument("--verbose", "-v", action="store_true", required=False,
                        help="Turn on verbose logging")
    args = parser.parse_args()
    if args.verbose:
        logger.setLevel(logging.DEBUG)
    logger.debug(f"Command Line arguments: {args}")
    logger.info(f"Proccessing Leader Map - `{args.leader_map}`")
    leader_map = get_leader_map(args.leader_map)
    logger.debug(f"Leader Map:\n{json.dumps(leader_map, indent=4)}")
    roster = {}
    for registrants in args.registrants:
        logger.info(f"Processing registrants from `{registrants}`")
        roster.update(load_registrants(registrants))
    logger.debug(f"Roster:\n{json.dumps(roster, indent=4)}")
    output = args.output_file or f"./Roster_{args.study_name.replace(' ', '_')}_{datetime.now().strftime('%m%d%y')}.xlsx"
    logger.info(f"Writing roster to `{output}`")
    writer = pd.ExcelWriter(output, engine="openpyxl", mode="w")
    logger.info(f"Generating At A Glance View")
    logger.debug(f"Roster keys: {roster.keys()}")
    logger.debug(f"leader_map keys: {leader_map.keys()}")
    at_a_glance(roster, leader_map, writer, num_chunks=args.at_a_glance_size)
    logger.info(f"Generating Leader Roster")
    leader_roster(leader_map, roster, writer)
    for group, groster in sorted(roster.items()):
        logger.info(f"Generating Roster for {group}")
        group_roster(group, groster, writer)
    writer.save()


if __name__ == '__main__':
    main()